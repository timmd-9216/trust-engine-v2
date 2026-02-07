#!/usr/bin/env python3
"""
Build analyze_corpus JSON input from deliveries Excel (e.g. proyecto_multitudes).

Reads data/_*.xlsx (or given path), maps columns to
the schema expected by POST /analyze-corpus, writes a JSON file and prints the
curl command to send it to the NLP service.

Schema produced:
  {
    "posts": [ {"full_text": "...", "user_screen_name": "...", "candidate_id": "..."}, ... ],
    "candidate_entities": [ "candidate_id1", ... ],  // unique candidate_id from data
    "top_negative_k": 20
  }

Sheets used:
  - keywordpost_yt: title + description -> full_text, channel_title -> user_screen_name, candidate_id
  - replies_yt:     text -> full_text, username -> user_screen_name, candidate_id
  - (ownpost_ig, ownpost_tw, replies_ig, replies_tw: same logic if columns exist and sheet has data)
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        "This script requires openpyxl. Install with: poetry add --group dev openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)


def _safe_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return s if s else ""


def _row_to_post_from_keywordpost(row: tuple, header: tuple) -> dict | None:
    """Map a row from keywordpost_yt (or similar) to one post."""
    by_col = dict(zip(header, row))
    title = _safe_str(by_col.get("title"))
    description = _safe_str(by_col.get("description"))
    full_text = f"{title} {description}".strip()
    if not full_text:
        return None
    user = _safe_str(by_col.get("channel_title") or by_col.get("username"))
    candidate_id = _safe_str(by_col.get("candidate_id"))
    return {
        "full_text": full_text,
        "user_screen_name": user or "unknown",
        "candidate_id": candidate_id or "unknown",
    }


def _row_to_post_from_replies(row: tuple, header: tuple) -> dict | None:
    """Map a row from replies_yt (or similar) to one post."""
    by_col = dict(zip(header, row))
    text = _safe_str(by_col.get("text"))
    if not text:
        return None
    user = _safe_str(by_col.get("username") or by_col.get("channel_title"))
    candidate_id = _safe_str(by_col.get("candidate_id"))
    return {
        "full_text": text,
        "user_screen_name": user or "unknown",
        "candidate_id": candidate_id or "unknown",
    }


def collect_posts_from_workbook(
    path: Path,
    *,
    max_posts: int | None = None,
    sheets_keyword: tuple[str, ...] = ("keywordpost_yt", "ownpost_tw", "ownpost_ig"),
    sheets_replies: tuple[str, ...] = ("replies_yt", "replies_tw", "replies_ig"),
) -> tuple[list[dict], set[str]]:
    """Load workbook and collect posts + unique candidate_ids."""
    wb = load_workbook(path, read_only=True)
    posts: list[dict] = []
    candidate_ids: set[str] = set()

    def add(post: dict | None) -> None:
        if post is None:
            return
        if max_posts is not None and len(posts) >= max_posts:
            return
        posts.append(post)
        if post.get("candidate_id") and post["candidate_id"] != "unknown":
            candidate_ids.add(post["candidate_id"])

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = tuple(_safe_str(c) for c in rows[0])
        if sheet_name in sheets_keyword:
            for row in rows[1:]:
                if max_posts is not None and len(posts) >= max_posts:
                    break
                post = _row_to_post_from_keywordpost(row, header)
                add(post)
        elif sheet_name in sheets_replies:
            for row in rows[1:]:
                if max_posts is not None and len(posts) >= max_posts:
                    break
                post = _row_to_post_from_replies(row, header)
                add(post)

    wb.close()
    return posts, candidate_ids


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build analyze_corpus JSON from deliveries Excel and print curl command.",
    )
    parser.add_argument(
        "xlsx",
        nargs="?",
        default="data/argentina.xlsx",
        help="Path to the Excel file",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="data/analyze_corpus_input.json",
        help="Output JSON path (default: data/analyze_corpus_input.json)",
    )
    parser.add_argument(
        "-n",
        "--max-posts",
        type=int,
        default=None,
        help="Max number of posts to include (default: all)",
    )
    parser.add_argument(
        "--top-negative-k",
        type=int,
        default=20,
        help="Value for top_negative_k (default: 20)",
    )
    parser.add_argument(
        "--url",
        default="http://127.0.0.1:8082/analyze-corpus",
        help="URL for analyze_corpus endpoint (default: http://127.0.0.1:8082/analyze-corpus)",
    )
    parser.add_argument(
        "--no-curl",
        action="store_true",
        help="Do not print the curl command",
    )
    args = parser.parse_args()

    path = Path(args.xlsx)
    if not path.exists():
        print(f"Error: File not found: {path}", file=sys.stderr)
        return 1

    posts, candidate_ids = collect_posts_from_workbook(path, max_posts=args.max_posts)
    if not posts:
        print("Error: No posts extracted from the Excel file.", file=sys.stderr)
        return 1

    payload = {
        "posts": posts,
        "candidate_entities": sorted(candidate_ids) if candidate_ids else [],
        "top_negative_k": args.top_negative_k,
    }

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"Wrote {len(posts)} posts to {out}")
    print(
        f"Unique candidate_id: {len(candidate_ids)} -> {sorted(candidate_ids)[:20]}{'...' if len(candidate_ids) > 20 else ''}"
    )

    if not args.no_curl:
        # Curl that POSTs the JSON file
        curl_cmd = (
            f'curl -s -X POST "{args.url}" \\\n'
            f'  -H "Content-Type: application/json" \\\n'
            f"  -d @{out}"
        )
        print("\n# Run this to call the analyze_corpus endpoint:\n")
        print(curl_cmd)
        print("\n# Pretty-print response (optional):\n")
        print(
            f'curl -s -X POST "{args.url}" -H "Content-Type: application/json" -d @{out} | python3 -m json.tool'
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
